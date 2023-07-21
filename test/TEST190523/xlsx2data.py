import openpyxl

def move_columns(filename):
    # Load the workbook
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active

    # Get the number of columns in the sheet
    num_columns = 10207

    # Iterate over columns starting from the third column
    for col in range(3, num_columns + 1):
        print(col)
        # Get the values from the column
        for cell in sheet.iter_cols(min_col=col, max_col=col, min_row=2):
            None
        values = [c.value for c in cell]

        # Insert the values under the second column
        sheet.insert_cols(2)
        for row, value in enumerate(values, start=2):
            sheet.cell(row=row, column=2).value = value

    # Repeat the first column under itself
    first_column_values = [cell.value for cell in sheet.iter_cols(min_col=1, max_col=1, min_row=2)]
    sheet.insert_cols(1)
    for row, value in enumerate(first_column_values, start=2):
        sheet.cell(row=row, column=1).value = value

    # Save the modified workbook
    modified_filename = f"modified_{filename}"
    workbook.save(modified_filename)
    print(f"Modified file saved as '{modified_filename}'")

# Usage example
filename = "datalv.xlsx"
move_columns(filename)
